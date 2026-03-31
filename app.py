import random
import math
import datetime
import calendar
import json
import holidays
import threading, uuid
from dateutil.parser import parse
import os
import io
import pandas as pd
from sqlalchemy import text
from flask_basicauth import BasicAuth
from dotenv import load_dotenv
from helper import *
from helper import save_prior_last_two as save_prior_last_two_db
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, g, session, send_file, abort
from flask_wtf import CSRFProtect
from flask_wtf.csrf import generate_csrf


def coerce_json_column(value):
    """
    Read JSON from a DB column that may be NULL or already decoded as dict
    (e.g. Postgres JSON/JSONB). Never call json.loads on None.
    """
    if value is None or value == "":
        return None
    if isinstance(value, dict):
        return value
    if isinstance(value, (str, bytes, bytearray)):
        try:
            return json.loads(value)
        except (TypeError, ValueError, json.JSONDecodeError):
            return None
    return None


def create_app():
    
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY","dev-key")
    # Cookie and request safety
    is_dev = os.environ.get("FLASK_ENV", "development") == "development"
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    app.config["SESSION_COOKIE_SECURE"] = not is_dev
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_CONTENT_LENGTH_BYTES", 2 * 1024 * 1024))
    # CSRF
    csrf = CSRFProtect(app)

    # login for stuff
    load_dotenv()
    basic_auth = BasicAuth(app)
    app.config['BASIC_AUTH_USERNAME'] = os.getenv('BASIC_AUTH_USERNAME')
    app.config['BASIC_AUTH_PASSWORD'] = os.getenv('BASIC_AUTH_PASSWORD')

    with app.app_context():
        init_db()

    solve_jobs = {}

    def convert_schedule_ids_to_names(schedule, id_to_name):
        """
        Convert a day/level schedule keyed by surgeon IDs (or names) into a
        name-based schedule using the provided lookup.
        """
        sched_by_name = {}
        for day, assigns in (schedule or {}).items():
            if not isinstance(assigns, dict):
                continue
            sched_by_name.setdefault(day, {})
            for lvl, val in assigns.items():
                if val in [None, ""]:
                    sched_by_name[day][lvl] = None
                    continue
                try:
                    sid = int(val)
                    sched_by_name[day][lvl] = id_to_name.get(sid)
                except Exception:
                    sched_by_name[day][lvl] = val
        return sched_by_name

    def normalize_schedule_for_template(sched):
        """
        Ensure schedule is a day -> {level: name} dict for new_schedule template/JS.
        Handles null/malformed per-day payloads and accidental error payloads in DB.
        """
        if sched is None:
            return None
        if not isinstance(sched, dict):
            return None
        # Solver/validation error objects saved or returned as dict
        if "errors" in sched and set(sched.keys()) <= {"errors", "analysis"}:
            return None
        out = {}
        for day, assigns in sched.items():
            try:
                datetime.date.fromisoformat(str(day))
            except Exception:
                continue
            if isinstance(assigns, dict):
                out[day] = assigns
            elif assigns is None:
                out[day] = {}
            else:
                out[day] = {}
        return out

    def extract_solver_mode(schedule_obj):
        if not isinstance(schedule_obj, dict):
            return None
        mode = schedule_obj.get("__solver_mode__")
        if mode in [None, ""]:
            return None
        return str(mode)

    @app.context_processor
    def utility_processor():
        return {
            'parse_call_levels': parse_call_levels,
            'csrf_token': lambda: generate_csrf()
        }
    
    @app.after_request
    def set_security_headers(resp):
        # Core hardening headers
        resp.headers.setdefault("X-Content-Type-Options", "nosniff")
        resp.headers.setdefault("X-Frame-Options", "DENY")
        resp.headers.setdefault("Referrer-Policy", "same-origin")
        # Relaxed but safer CSP to avoid breaking CDN and inline usage
        csp = (
            "default-src 'self'; "
            "base-uri 'self'; object-src 'none'; frame-ancestors 'none'; "
            "img-src 'self' data:; style-src 'self' 'unsafe-inline' https:; "
            "script-src 'self' 'unsafe-inline' https:;"
        )
        resp.headers.setdefault("Content-Security-Policy", csp)
        # HSTS only when HTTPS or explicitly enabled
        if request.is_secure or os.environ.get("ENABLE_HSTS") == "1":
            resp.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
        return resp
    
    @app.teardown_appcontext
    def teardown_db(exception):
        close_db(exception)
            
    #############################################
    # Surgeon Management Endpoints
    #############################################

    @app.route('/surgeons')
    @basic_auth.required
    def list_surgeons():
        surgeons = get_all_surgeons()
        if request.args.get('sort'):
            level_order = {"1A": 1, "1B": 1, "2A": 2, "2B": 2, "3": 3, "4": 4}
            def get_lowest_level(s):
                levels = parse_call_levels(s.get("call_levels", ""))
                if not levels:
                    return 99
                orders = [level_order.get(l, 99) for l in levels]
                return min(orders)
            surgeons.sort(key=lambda s: (get_lowest_level(s), s["name"].lower()))
        return render_template('surgeons_list.html', surgeons=surgeons)
    
    @app.route('/surgeons/sort_by_team_then_level')
    @basic_auth.required
    def list_surgeons_by_team_then_level():
        surgeons = get_all_surgeons()

        # same order you use elsewhere for levels
        level_order = {"1A":1, "1B":1, "2A":2, "2B":2, "3":3, "4":4}
        def lowest_level(s):
            levels = parse_call_levels(s.get("call_levels",""))
            if not levels:
                return float('inf')
            return min(level_order.get(l, float('inf')) for l in levels)

        # sort by: team (None last), then by lowest call level, then name
        surgeons.sort(key=lambda s: (
            (s['team'] or 'zzz'),
            lowest_level(s),
            s['name'].lower()
        ))

        return render_template('surgeons_list.html', surgeons=surgeons)

    @app.route('/surgeons/add', methods=['GET', 'POST'])
    @basic_auth.required
    def add_surgeon():
        if request.method == 'POST':
            name = request.form['name']
            call_levels_list = request.form.getlist('call_levels')
            call_levels = ','.join(call_levels_list)
            team  = request.form['team']
            try:
                less_credit = int(request.form.get('manual_less_calls_credit', 0))
            except Exception:
                less_credit = 0
            try:
                more_credit = int(request.form.get('manual_more_calls_credit', 0))
            except Exception:
                more_credit = 0
            less_credit = max(0, less_credit)
            more_credit = max(0, more_credit)
            db = get_db()
            with db.begin():
                db.execute(
                    text(
                        """
                        INSERT INTO surgeons
                            (name, call_levels, team, manual_less_calls_credit, manual_more_calls_credit)
                        VALUES
                            (:name, :levels, :team, :less_credit, :more_credit)
                        """
                    ),
                    {
                        "name": name,
                        "levels": call_levels,
                        "team": team,
                        "less_credit": less_credit,
                        "more_credit": more_credit,
                    }
                )
            flash("Surgeon added successfully!")
            return redirect(url_for('list_surgeons'))
    # Pass a default surgeon dict with "call_levels" defined so the template doesn't error out.
        default_surgeon = {
            "name": "",
            "call_levels": "",
            "team": "",
            "id": 0,
            "manual_less_calls_credit": 0,
            "manual_more_calls_credit": 0,
        }
        return render_template('surgeon_form.html', surgeon=default_surgeon, action="Add")

    @app.route('/surgeons/add_quick', methods=['POST'])
    @basic_auth.required
    def add_surgeon_quick():
        team = (request.form.get('team') or '').strip()
        if team == "__UNASSIGNED__":
            team = ""
        db = get_db()
        with db.begin():
            db.execute(
                text(
                    """
                    INSERT INTO surgeons (name, call_levels, nlth, team, manual_less_calls_credit, manual_more_calls_credit)
                    VALUES (:name, :levels, :nlth, :team, :less_credit, :more_credit)
                    """
                ),
                {
                    "name": "",
                    "levels": "",
                    "nlth": False,
                    "team": team,
                    "less_credit": 0,
                    "more_credit": 0,
                }
            )
        flash(f"Added a new blank surgeon to {(team or 'Unassigned')}.", "success")
        return redirect(url_for('list_surgeons'))
    @app.route('/surgeons/edit/<int:surgeon_id>', methods=['GET', 'POST'])
    @basic_auth.required
    def edit_surgeon(surgeon_id):
        db = get_db()
        if request.method == 'POST':
            try:
                name = request.form.get('name', '').strip()
                call_levels_list = request.form.getlist('call_levels')
                call_levels = ','.join(call_levels_list)
                team = request.form.get('team', '')
                try:
                    less_credit = int(request.form.get('manual_less_calls_credit', 0))
                except Exception:
                    less_credit = 0
                try:
                    more_credit = int(request.form.get('manual_more_calls_credit', 0))
                except Exception:
                    more_credit = 0
                less_credit = max(0, less_credit)
                more_credit = max(0, more_credit)
                with db.begin():
                    db.execute(
                        text(
                            """
                            UPDATE surgeons
                            SET name = :name,
                                call_levels = :levels,
                                team = :team,
                                manual_less_calls_credit = :less_credit,
                                manual_more_calls_credit = :more_credit
                            WHERE id = :id
                            """
                        ),
                        {
                            "name": name,
                            "levels": call_levels,
                            "team": team,
                            "less_credit": less_credit,
                            "more_credit": more_credit,
                            "id": surgeon_id,
                        }
                    )
                flash("Surgeon updated successfully!")
                return redirect(url_for('list_surgeons'))
            except Exception as e:
                import traceback
                traceback.print_exc()
                flash(f"Failed to update surgeon: {e}", "error")
                return redirect(url_for('edit_surgeon', surgeon_id=surgeon_id))
        # GET path
        row = db.execute(
            text("SELECT * FROM surgeons WHERE id = :id"),
            {"id": surgeon_id}
        ).mappings().fetchone()
        if not row:
            flash("Surgeon not found!")
            return redirect(url_for('list_surgeons'))
        surgeon = dict(row)
        return render_template('surgeon_form.html', surgeon=surgeon, action="Edit")

    @app.route('/surgeons/update/<int:surgeon_id>', methods=['POST'])
    @basic_auth.required
    def update_surgeon_inline(surgeon_id):
        name = request.form['name']
        call_levels_list = request.form.getlist('call_levels')
        call_levels = ','.join(call_levels_list)
        team = request.form['team']
        try:
            less_credit = int(request.form.get('manual_less_calls_credit', 0))
        except Exception:
            less_credit = 0
        try:
            more_credit = int(request.form.get('manual_more_calls_credit', 0))
        except Exception:
            more_credit = 0
        less_credit = max(0, less_credit)
        more_credit = max(0, more_credit)

        db = get_db()
        with db.begin():
            db.execute(
                    text(
                        """
                        UPDATE surgeons
                        SET name = :name,
                            call_levels = :levels,
                            team = :team,
                            manual_less_calls_credit = :less_credit,
                            manual_more_calls_credit = :more_credit
                        WHERE id = :id
                        """
                    ),
                    {
                        "name": name,
                        "levels": call_levels,
                        "team": team,
                        "less_credit": less_credit,
                        "more_credit": more_credit,
                        "id": surgeon_id,
                    }
            )
            flash("Surgeon updated successfully!")
            return redirect(url_for('list_surgeons'))

    @app.route('/update_all_surgeons', methods=['POST'])
    @basic_auth.required
    def update_all_surgeons():
        db = get_db()
        # Prepare a single text‐Clause for performance
        stmt = text("""
            UPDATE surgeons
            SET name       = :name,
                call_levels= :levels,
                nlth       = :nlth,
                team       = :team
            WHERE id         = :id
        """)

        try:
            with db.begin():
                # Fetch within the same transaction to avoid nested begin/autobegin conflicts
                surgeons = db.execute(text("SELECT * FROM surgeons")).mappings().all()
                for surgeon in surgeons:
                    sid = surgeon['id']
                    name_field = f"name_{sid}"
                    levels_field = f"call_levels_{sid}"
                    nlth_field = f"nlth_{sid}"
                    team_field = f"team_{sid}"

                    new_name   = request.form.get(name_field, surgeon['name'])
                    new_levels = request.form.getlist(levels_field)
                    new_levels_str = ",".join(new_levels)
                    new_team = request.form.get(team_field) or ""

                    # Checkbox: if the field is present in form data, checkbox was checked
                    new_nlth = (request.form.get(nlth_field) == 'on')

                    db.execute(
                        stmt,
                        {
                            "name":   new_name,
                            "levels": new_levels_str,
                            "nlth":   new_nlth,
                            "team":   new_team,
                            "id":     sid
                        }
                    )
            flash("Surgeon updates applied successfully!")
        except Exception as e:
            import traceback
            traceback.print_exc()
            flash(f"Failed to update surgeons: {e}", "error")
        return redirect(url_for('list_surgeons'))

    #############################################
    # Maximum Calls Configuration Endpoint
    #############################################

    @app.route('/config_max_calls', methods=['GET', 'POST'])
    @basic_auth.required
    def config_max_calls():
        if request.method == 'POST':
            new_config = {
                "1": int(request.form.get("group_1", 10)),
                "2": int(request.form.get("group_2", 10)),
                "3": int(request.form.get("group_3", 10)),
                "4": int(request.form.get("group_4", 10))
            }
            update_max_calls_config(new_config)
            flash("Maximum calls configuration updated successfully!")
            return redirect(url_for('config_max_calls'))
        config = get_max_calls_config()
        return render_template('config_max_calls.html', config=config)

    #############################################
    # Global Config for No Call Request Constraint Endpoint
    #############################################

    @app.route('/global_config', methods=['GET', 'POST'])
    @basic_auth.required
    def global_config_page():
        teams = ['Team 1','Team 2','Team 3','Team 4','Urology']
        if request.method == 'POST':
            existing_config = get_global_config()
            days = list(range(7))
            new_prefs = {}
            for team in teams:
                new_prefs[team] = {
                wd: int(request.form.get(f'pref_{team}_{wd}', 0))
                for wd in days
                }
            update_team_day_prefs(new_prefs)
            no_call_hard_val = request.form.get("no_call_hard", "1")
            pre_unavail_mode = request.form.get("pre_unavail_mode", "soft") or "soft"
            if pre_unavail_mode not in ("hard", "soft", "off"):
                pre_unavail_mode = "soft"
            # sliders
            fairness_weight = request.form.get("fairness_weight", "1000")
            gamma_no_call = request.form.get("gamma_no_call", "10")
            gamma_unavail_prev = request.form.get("gamma_unavail_prev", "5")
            gamma_spacing = request.form.get("gamma_spacing", "10")
            spacing_threshold = request.form.get("spacing_threshold", "7")
            gamma_team_pref = request.form.get("gamma_team_pref", "10")
            gamma_weekend_balance = request.form.get("gamma_weekend_balance", "50")
            gamma_consec_weekend = request.form.get("gamma_consec_weekend", "20")
            gamma_weekend_team_diversity = request.form.get("gamma_weekend_team_diversity", "50")
            gamma_balance = request.form.get("gamma_balance", "100")
            max_weekend_calls = request.form.get("max_weekend_calls", "3")
            min_calls_nlth = request.form.get("min_calls_nlth", "3")
            max_calls_level1 = request.form.get("max_calls_level1", "10")
            # unavailability credit controls
            gamma_unavail_credit = request.form.get('gamma_unavail_credit', '50')
            unavail_credit_days = request.form.get('unavail_credit_days', '7')
            # deadline controls
            unavail_deadline_day = request.form.get("unavail_deadline_day", "20")
            unavail_deadline_time = request.form.get("unavail_deadline_time", "23:59")
            fairness_fallback_policy = request.form.get("fairness_fallback_policy", "auto_relax")
            if fairness_fallback_policy not in ("auto_relax", "no_fallback"):
                fairness_fallback_policy = "auto_relax"
            # checkboxes
            def cb(name, default='0'):
                """
                Preserve existing persisted value when a checkbox field is absent.
                This prevents partial/legacy forms from silently disabling flags.
                """
                if name in request.form:
                    return '1' if request.form.get(name) == '1' else '0'
                return str(existing_config.get(name, default))
            flags = {
                "enable_spacing_penalty": cb("enable_spacing_penalty", "1"),
                "enable_availability_unavail_prev_penalty": cb("enable_availability_unavail_prev_penalty", "1"),
                "enable_availability_nocall_penalty": cb("enable_availability_nocall_penalty", "1"),
                "enable_unavail_credit": cb("enable_unavail_credit", "1"),
                "enable_force_1B_weekend": cb("enable_force_1B_weekend", "1"),
                "enable_level2_supervision": cb("enable_level2_supervision", "1"),
                "enable_group4_2B3_ban": cb("enable_group4_2B3_ban", "1"),
                "enable_max_2B_group4": cb("enable_max_2B_group4", "1"),
                "enable_max_calls_level1": cb("enable_max_calls_level1", "1"),
                "enable_nlth_rules": cb("enable_nlth_rules", "1"),
                "enable_weekend_balance": cb("enable_weekend_balance", "1"),
                "enable_weekend_consecutive_penalty": cb("enable_weekend_consecutive_penalty", "1"),
                "enable_weekend_team_diversity_enable": cb("enable_weekend_team_diversity_enable", "1"),
                "enable_team_day_prefs": cb("enable_team_day_prefs", "1"),
                "enable_horizon_fairness": cb("enable_horizon_fairness", "0"),
                "fairness_cap_uses_credit": cb("fairness_cap_uses_credit", "0"),
                "enable_fairness_hard_cap": cb("enable_fairness_hard_cap", "1"),
                "enable_two_pass_fairness_priority": cb("enable_two_pass_fairness_priority", "1"),
                "solver_debug": cb("solver_debug", "0"),
                "enable_unavail_deadline": cb("enable_unavail_deadline", "0"),
            }
            update_global_config({
                "no_call_hard": no_call_hard_val,
                "pre_unavail_mode": pre_unavail_mode,
                "fairness_weight": fairness_weight,
                "gamma_no_call": gamma_no_call,
                "gamma_unavail_prev": gamma_unavail_prev,
                "gamma_spacing": gamma_spacing,
                "spacing_threshold": spacing_threshold,
                "gamma_team_pref": gamma_team_pref,
                "gamma_weekend_balance": gamma_weekend_balance,
                "gamma_consec_weekend": gamma_consec_weekend,
                "gamma_weekend_team_diversity": gamma_weekend_team_diversity,
                "gamma_balance": gamma_balance,
                "gamma_unavail_credit": gamma_unavail_credit,
                "unavail_credit_days": unavail_credit_days,
                "max_weekend_calls": max_weekend_calls,
                "min_calls_nlth": min_calls_nlth,
                "max_calls_level1": max_calls_level1,
                "gamma_2b_usage": request.form.get("gamma_2b_usage", "0"),
                "fairness_hard_cap_range": request.form.get("fairness_hard_cap_range", "1"),
                "fairness_fallback_policy": fairness_fallback_policy,
                "unavail_deadline_day": unavail_deadline_day,
                "unavail_deadline_time": unavail_deadline_time,
                **flags
            })
            flash("Global configuration saved.", "success")
            return redirect(url_for('global_config_page'))
        else:
            config = get_global_config()
            prefs  = get_team_day_prefs()
            return render_template('global_config.html', config=config, prefs=prefs, teams=teams)

            
    #############################################
    # Constraint weights Endpoints
    #############################################

    @app.route('/constraint_weights', methods=['GET', 'POST'])
    @basic_auth.required
    def constraint_weights():
        if request.method == 'POST':
            # Retrieve new weight values from the form.
            fairness_weight = request.form.get('fairness_weight', '1000')
            gamma_no_call = request.form.get('gamma_no_call', '10')
            gamma_unavail_prev = request.form.get('gamma_unavail_prev', '5')
            gamma_1B = request.form.get('gamma_1B', '1')
            gamma_spacing = request.form.get('gamma_spacing', '10')
            spacing_threshold = request.form.get("spacing_threshold", "7")
            gamma_team_pref = request.form.get("gamma_team_pref", "10")
            gamma_weekend_team_diversity = request.form.get("gamma_weekend_team_diversity", "50")
            gamma_unavail_credit = request.form.get('gamma_unavail_credit', '50')
            unavail_credit_days = request.form.get('unavail_credit_days', '7')
            # Feature flags via checkboxes (unchecked -> missing -> treat as '0')
            def cb(name):
                return '1' if request.form.get(name) == '1' else '0'
            flags = {
                "enable_force_1B_weekend": cb("enable_force_1B_weekend"),
                "enable_level2_supervision": cb("enable_level2_supervision"),
                "enable_group4_2B3_ban": cb("enable_group4_2B3_ban"),
                "enable_max_2B_group4": cb("enable_max_2B_group4"),
                "enable_max_calls_level1": cb("enable_max_calls_level1"),
                "enable_nlth_rules": cb("enable_nlth_rules"),
                "enable_weekend_consecutive_penalty": cb("enable_weekend_consecutive_penalty"),
                "enable_weekend_balance": cb("enable_weekend_balance"),
                "enable_weekend_team_diversity_enable": cb("enable_weekend_team_diversity_enable"),
                "enable_team_day_prefs": cb("enable_team_day_prefs"),
                "enable_availability_unavail_prev_penalty": cb("enable_availability_unavail_prev_penalty"),
                "enable_availability_nocall_penalty": cb("enable_availability_nocall_penalty"),
                "enable_spacing_penalty": cb("enable_spacing_penalty"),
                "enable_fairness_diff_all": cb("enable_fairness_diff_all"),
                "enable_deviation_sum": cb("enable_deviation_sum"),
            }
            pre_unavail_mode = request.form.get("pre_unavail_mode", "soft") or "soft"
            if pre_unavail_mode not in ("hard", "soft", "off"):
                pre_unavail_mode = "soft"
            # Update global configuration.
            update_global_config({
                "fairness_weight": fairness_weight,
                "gamma_no_call": gamma_no_call,
                "gamma_unavail_prev": gamma_unavail_prev,
                "gamma_1B": gamma_1B,
                "gamma_spacing": gamma_spacing,
                "spacing_threshold": spacing_threshold,
                "gamma_team_pref": gamma_team_pref,
                "gamma_weekend_team_diversity": gamma_weekend_team_diversity,
                "gamma_unavail_credit": gamma_unavail_credit,
                "unavail_credit_days": unavail_credit_days,
                "pre_unavail_mode": pre_unavail_mode,
                **flags
            })
            flash("Constraint weights updated successfully!")
            return redirect(url_for('constraint_weights'))
        else:
            config = get_global_config()
            return render_template('constraint_weights.html', config=config)


    #############################################
    # Schedule Generation and Saving Endpoints
    #############################################

    # (About page removed as part of revert)
    def run_solver_job(job_id, days, surgeons, prev_schedule, public_holidays, preassignments, time_limit_seconds: int = 30, allow_empty: bool = False, horizon_prior=None):
        # we only import the solver function here—never re‑import `app` or `solve_jobs`
        from scheduler import solve_schedule_or_tools

        # mark the job as running
        solve_jobs[job_id] = {
            'status': 'running',
            'best':   None,
            'cancel': False,
            'solution': None,
            'solver_mode': None,
            'public_holidays': sorted(list(public_holidays or [])),
        }

        # we need the app_context so that any DB or flask helpers will work
        with app.app_context():
            sched, cost = solve_schedule_or_tools(
                days, surgeons, prev_schedule, public_holidays, preassignments,
                time_limit_seconds=time_limit_seconds,
                allow_empty=allow_empty,
                horizon_prior_counts=horizon_prior
            )
            solver_mode = extract_solver_mode(sched)
            clean_sched = normalize_schedule_for_template(sched) if isinstance(sched, dict) else sched

            if clean_sched is not None and not (isinstance(clean_sched, dict) and 'errors' in clean_sched):
                solve_jobs[job_id].update({
                    'status':   'done',
                    'solution': clean_sched,
                    'best':     cost,
                    'solver_mode': solver_mode,
                })
            else:
                solve_jobs[job_id].update({
                    'status': 'failed',
                    'solution': sched,
                    'solver_mode': solver_mode,
                })

    @app.route('/new_schedule', methods=['GET'])
    @basic_auth.required
    def new_schedule():
        import datetime, calendar, json
        from scheduler import solve_schedule_or_tools

        # ── 1) Ensure year/month are always in the URL ──
        if not request.args.get('year') or not request.args.get('month'):
            today = datetime.date.today()
            next_year = today.year + 1 if today.month == 12 else today.year
            next_month = 1 if today.month == 12 else today.month + 1
            return redirect(url_for(
                'new_schedule',
                year=next_year,
                month=next_month
            ))

        # ── 2) Parse year & month ──
        year_sel, month_sel = get_year_month()

        # ── 3) Build list of days for that month ──
        days_sel = [
            datetime.date(year_sel, month_sel, d).isoformat()
            for d in range(1, calendar.monthrange(year_sel, month_sel)[1] + 1)
        ]

        # ── 4) HK holidays ──
        hk_h = holidays.HK(years=[year_sel])
        public_holidays = {
            d.isoformat() for d in hk_h 
            if d.year == year_sel and d.month == month_sel
        }

        # ── 5) Compute previous month references (for labels/UI) ──
        if month_sel == 1:
            prev_year, prev_month = year_sel - 1, 12
        else:
            prev_year, prev_month = year_sel, month_sel - 1
        db = get_db()
        # compute the last two dates of previous month for display/overlay
        prev_month_num_days = calendar.monthrange(prev_year, prev_month)[1]
        prev_day_minus_2 = datetime.date(prev_year, prev_month, prev_month_num_days - 1).isoformat()
        prev_day_minus_1 = datetime.date(prev_year, prev_month, prev_month_num_days).isoformat()

        # ── Load preassignments for the current month ──
        stmt_pre = text(
            "SELECT preassignment_data FROM preassignments WHERE year = :year AND month = :month"
        )
        result_pre = db.execute(stmt_pre, {"year": year_sel, "month": month_sel})
        row_pre = result_pre.mappings().fetchone()
        # Avoid leaking DB details in logs in production
        if app.debug:
            print("Fetched preassignment row:", row_pre)

        if row_pre and row_pre['preassignment_data'] is not None and row_pre['preassignment_data'] != "":
            raw_pre = row_pre['preassignment_data']
            if app.debug:
                print("Raw preassignment data (raw_pre):", raw_pre)
            try:
                preassignments = raw_pre if isinstance(raw_pre, dict) else json.loads(raw_pre)
            except Exception as e:
                if app.debug:
                    print("Error decoding preassignments:", e)
                preassignments = {}
        else:
            preassignments = {}

        if app.debug:
            print("Loaded preassignments:", preassignments)

        # ── 6) Do we already have one in the DB? ──
        row = db.execute(
            text("SELECT * FROM saved_schedule WHERE year = :y AND month = :m"),
            {"y": year_sel, "m": month_sel}
        ).fetchone()

        # Build candidates per level for UI selects
        surgeons = get_all_surgeons()
        def candidates_for(level: str):
            return [s for s in surgeons if level in parse_call_levels(s.get("call_levels", ""))]
        candidates = {lvl: sorted(candidates_for(lvl), key=lambda s: s["name"]) for lvl in ["1A","1B","2A","2B","3","4"]}
        cohort_summary = build_half_year_cohort_summary(year_sel, month_sel, surgeons=surgeons)
        solver_mode_used = None

        generate_flag = request.args.get('generate')
        if generate_flag:
            # ▶︎ Preview only; do *not* save
            # Build prev_schedule ONLY from saved prior-last-two for pruning
            saved_prior = get_prior_last_two(year_sel, month_sel)
            id_to_name = {s['id']: s['name'] for s in surgeons}
            prev_from_prior = {}
            if saved_prior.get('m2'):
                prev_from_prior[prev_day_minus_2] = {lev: id_to_name.get(int(sid)) for lev, sid in saved_prior['m2'].items() if id_to_name.get(int(sid))}
            if saved_prior.get('m1'):
                prev_from_prior[prev_day_minus_1] = {lev: id_to_name.get(int(sid)) for lev, sid in saved_prior['m1'].items() if id_to_name.get(int(sid))}

            sched, cost = solve_schedule_or_tools(
                days_sel,
                surgeons,
                prev_schedule=prev_from_prior or None,
                public_holidays=public_holidays,
                preassignments=preassignments
            )
            if sched is None:
                flash("No feasible schedule found.", "error")
                return render_template('no_schedule.html')
            solver_mode_used = extract_solver_mode(sched)
            sched = normalize_schedule_for_template(sched)
        else:
            # ▶︎ No preview request → show the saved one (if any)
            cost = None

            # Fetch as a mapping so row['schedule_data'] works
            stmt = text(
                "SELECT schedule_data FROM saved_schedule "
                "WHERE year = :year AND month = :month"
            )
            result = db.execute(stmt, {"year": year_sel, "month": month_sel})
            row = result.mappings().fetchone()

            if row:
                sched = coerce_json_column(row['schedule_data'])
            else:
                sched = None
            sched = normalize_schedule_for_template(sched)

        # ── 7) Compute weekends set ──
        weekend_set = {
            d for d in days_sel
            if datetime.date.fromisoformat(d).weekday() >= 5
        }

        # ── 8) Render in both cases ──
        # load any saved prior-last-two for this (year, month)
        saved_prior = get_prior_last_two(year_sel, month_sel)
        return render_template(
            'new_schedule.html',
            schedule=sched,
            cost=cost,
            weekend_set=weekend_set,
            public_holidays=public_holidays,
            year=year_sel,
            month=month_sel,
            candidates=candidates,
            prev_day_minus_2=prev_day_minus_2,
            prev_day_minus_1=prev_day_minus_1,
            saved_prior=saved_prior,
            cohort_summary=cohort_summary,
            surgeons_meta=surgeons,
            solver_mode_used=solver_mode_used,
        )

    @app.route('/save_schedule', methods=['POST'])
    @basic_auth.required
    def save_schedule():
        # 1) Figure out which month/year we’re saving
        year_str = request.form.get('year') or request.args.get('year')
        month_str = request.form.get('month') or request.args.get('month')
        try:
            year = int(year_str)
            month = int(month_str)
        except (TypeError, ValueError):
            flash("Invalid year/month for saving schedule.", "error")
            # Redirect back to the new_schedule view, preserving whatever we have
            return redirect(url_for('new_schedule',
                                    year=year_str or None,
                                    month=month_str or None))

        # 2) Grab the JSON payload
        data = request.form.get('schedule_json')
        if not data:
            flash("No schedule data provided to save.", "error")
            return redirect(url_for('new_schedule', year=year, month=month))

        db = get_db()
        # Perform existence check and write within a single transaction so changes commit
        with db.begin():
            exists = db.execute(
                text("SELECT 1 FROM saved_schedule WHERE year = :y AND month = :m"),
                {"y": year, "m": month}
            ).fetchone()

            if exists:
                db.execute(
                    text(
                        "UPDATE saved_schedule "
                        f"SET schedule_data = {json_cast('sched')}, date_saved = {sql_now()} "
                        "WHERE year = :y AND month = :m"
                    ),
                    {"sched": data, "y": year, "m": month}
                )
            else:
                db.execute(
                    text(
                        "INSERT INTO saved_schedule "
                        "(year, month, schedule_data, date_saved) "
                        f"VALUES (:y, :m, {json_cast('sched')}, {sql_now()})"
                    ),
                    {"y": year, "m": month, "sched": data}
                )
        # Also save as a new version in versioned table
        db.commit()
        with db.begin():
            ver_row = db.execute(text("SELECT COALESCE(MAX(version),0) AS maxv FROM saved_schedule_versions WHERE year=:y AND month=:m"), {"y": year, "m": month}).mappings().fetchone()
            next_ver = (ver_row['maxv'] or 0) + 1
            db.execute(
                text(
                    f"INSERT INTO saved_schedule_versions (year, month, version, version_name, schedule_data, published) "
                    f"VALUES (:y, :m, :v, :vn, {json_cast('d')}, {sql_false()})"
                ),
                {"y": year, "m": month, "v": next_ver, "vn": f"v{next_ver}", "d": data}
            )
        flash("Schedule saved.", "success")
        if request.form.get("after_save") == "edit_publish":
            return redirect(url_for('edit_publish', year=year, month=month))
        return redirect(url_for('new_schedule', year=year, month=month))

    @app.route('/saved_schedule', methods=['GET'])
    def saved_schedule():
        year_sel, month_sel = get_year_month()
        # Generate the days for the selected period.
        days_sel = [datetime.date(year_sel, month_sel, d).isoformat() 
                    for d in range(1, calendar.monthrange(year_sel, month_sel)[1] + 1)]
        
        db = get_db()
        # Prefer published versioned schedule if available
        row_pub = db.execute(
            text(f"SELECT schedule_data FROM saved_schedule_versions WHERE year=:y AND month=:m AND published={sql_true()} ORDER BY version DESC LIMIT 1"),
            {"y": year_sel, "m": month_sel}
        ).mappings().fetchone()
        if row_pub:
            sched = coerce_json_column(row_pub['schedule_data']) or {}
        else:
            # Fallback: show message and empty content (do not show unpublished)
            flash("No published schedule for the selected month.", "warning")
            sched = {}

        # Use the locally computed days_sel to determine weekend_set and public holidays.
        weekend_set = {d for d in days_sel if datetime.date.fromisoformat(d).weekday() >= 5}
        hk_h = holidays.HK(years=[year_sel])
        public_holidays = {d.isoformat() for d in hk_h if d.year == year_sel and d.month == month_sel}
        return render_template('saved_schedule.html', schedule=sched, weekend_set=weekend_set, public_holidays=public_holidays,
                            year=year_sel, month=month_sel)

    @app.route('/save_prior_last_two', methods=['POST'])
    @basic_auth.required
    def save_prior_last_two():
        data = request.get_json()
        try:
            year = int(data.get('year'))
            month = int(data.get('month'))
        except Exception:
            return jsonify({'error': 'Invalid year/month'}), 400
        m2 = data.get('m2') or {}
        m1 = data.get('m1') or {}
        # sanitize: keys are levels, values are ints
        try:
            m2 = {k: int(v) for k, v in m2.items()}
            m1 = {k: int(v) for k, v in m1.items()}
        except Exception:
            return jsonify({'error': 'Invalid payload'}), 400
        save_prior_last_two_db(year, month, m2, m1)
        return jsonify({'ok': True})

    @app.route('/save_surgeon_call_credits', methods=['POST'])
    @basic_auth.required
    def save_surgeon_call_credits():
        data = request.get_json() or {}
        raw_credits = data.get("credits") or {}
        surgeons = get_all_surgeons()
        valid_ids = {int(s["id"]) for s in surgeons if s.get("id") is not None}
        payload = {}
        for sid_raw, row in raw_credits.items():
            try:
                sid = int(sid_raw)
            except Exception:
                continue
            if sid not in valid_ids:
                continue
            row = row or {}

            # New semantics: signed value (negative=fewer calls, positive=more calls).
            if "call_credit" in row:
                try:
                    signed_credit = int(row.get("call_credit", 0))
                except Exception:
                    return jsonify({"error": f"Invalid call_credit for surgeon {sid}"}), 400
                less_credit = max(0, -signed_credit)
                more_credit = max(0, signed_credit)
            else:
                # Backward compatibility with older split payload shape.
                try:
                    less_credit = int(row.get("less_calls_credit", 0))
                except Exception:
                    return jsonify({"error": f"Invalid less_calls_credit for surgeon {sid}"}), 400
                try:
                    more_credit = int(row.get("more_calls_credit", 0))
                except Exception:
                    return jsonify({"error": f"Invalid more_calls_credit for surgeon {sid}"}), 400
                if less_credit < 0 or more_credit < 0:
                    return jsonify({"error": f"Credits must be non-negative for surgeon {sid}"}), 400
            payload[sid] = {
                "less_calls_credit": less_credit,
                "more_calls_credit": more_credit,
            }

        update_surgeon_manual_call_credits(payload)
        return jsonify({"ok": True, "updated": len(payload)})

    @app.route('/export_schedule', methods=['POST'])
    @basic_auth.required
    def export_schedule():
        year = int(request.form['year'])
        month = int(request.form['month'])
        sched_json = request.form.get('schedule_json')
        if not sched_json:
            flash("Nothing to export!", "error")
            return redirect(url_for('new_schedule', year=year, month=month))

        # 1) Load schedule dict
        schedule = json.loads(sched_json)

        # 2) Build schedule DataFrame
        df_sched = (
            pd.DataFrame.from_dict(schedule, orient='index')
            .rename_axis('Date')
            .reset_index()
        )
        cols = ['Date','1A','1B','2A','2B','3','4']
        df_sched = df_sched[cols]

        # 3) Compute per-surgeon stats
        level_ranks = {"1A":1,"1B":1,"2A":2,"2B":3,"3":4,"4":5}
        stats = {}
        weekend = set(d for d in schedule if datetime.date.fromisoformat(d).weekday() >= 5)
        for date, assigns in schedule.items():
            is_wk = date in weekend
            for lvl, name in assigns.items():
                if not name:
                    continue
                rec = stats.setdefault(name, {"total_calls":0,"weekend_calls":0,"min_level_rank":None})
                rec["total_calls"] += 1
                if is_wk:
                    rec["weekend_calls"] += 1
                rank = level_ranks.get(lvl,99)
                if rec["min_level_rank"] is None or rank < rec["min_level_rank"]:
                    rec["min_level_rank"] = rank
        df_stats = (
            pd.DataFrame([
                {"Surgeon": name,
                 "Total Calls": v["total_calls"],
                 "Weekend Calls": v["weekend_calls"],
                 "Min Level Rank": v["min_level_rank"]}
                for name,v in stats.items()
            ])
            .sort_values(["Min Level Rank","Surgeon"])
            .reset_index(drop=True)
        )

        # 3b) Compute Half-Year (H1/H2-to-date including current month) per-group totals
        df_hy_stats = None
        try:
            from helper import get_published_schedule_version, get_all_surgeons, parse_call_levels, get_level2_group
            surgeons_list = get_all_surgeons()
            # Determine half-year months
            if 1 <= month <= 6:
                hy_months = list(range(1, month + 1))
            else:
                hy_months = list(range(7, month + 1))
            # Aggregate level counts per surgeon name
            levels = ["1A","1B","2A","2B","3","4"]
            hy_level_counts = {}
            def add_count(nm, lvl):
                if not nm:
                    return
                rec = hy_level_counts.setdefault(nm, {L:0 for L in levels})
                if lvl in rec:
                    rec[lvl] += 1
            # Prior months (published schedules)
            for m in [m for m in hy_months if m != month]:
                prev = get_published_schedule_version(year, m)
                if not prev:
                    continue
                for _, assigns in (prev or {}).items():
                    if not isinstance(assigns, dict):
                        continue
                    for L in levels:
                        add_count(assigns.get(L), L)
            # Current month (use in-memory schedule)
            for _, assigns in schedule.items():
                for L in levels:
                    add_count(assigns.get(L), L)

            # Group membership by surgeon name
            def has_level(s, L):
                return L in parse_call_levels(s.get('call_levels',''))
            group1_names = sorted({s['name'] for s in surgeons_list if (has_level(s,'1A') or has_level(s,'1B'))})
            l2_union_names = sorted({s['name'] for s in surgeons_list if get_level2_group(s) in (1,2,3)})
            group4_level_names = sorted({s['name'] for s in surgeons_list if has_level(s,'4')})
            group4_l2_names = sorted({s['name'] for s in surgeons_list if get_level2_group(s) == 4})
            s3_union_names = sorted({s['name'] for s in surgeons_list if has_level(s,'3') or s['name'] in group4_l2_names})

            # Per-surgeon half-year totals per fairness group
            g1_total = {}
            g2_total = {}
            g3_total = {}
            g4_total = {}
            for nm, lc in hy_level_counts.items():
                g1_total[nm] = int(lc.get('1A',0) + lc.get('1B',0))
                g2_total[nm] = int(lc.get('2A',0) + lc.get('2B',0))
                # group 3: level 3 for all; plus 2B only for grp4 surgeons
                g3_extra_2b = lc.get('2B',0) if nm in group4_l2_names else 0
                g3_total[nm] = int(lc.get('3',0) + g3_extra_2b)
                g4_total[nm] = int(lc.get('4',0))

            # Group averages (consider only cohort members)
            def avg(vals):
                vals = [v for v in vals if isinstance(v, (int, float))]
                return (sum(vals)/len(vals)) if vals else 0.0
            avg_g1 = avg([g1_total.get(nm,0) for nm in group1_names])
            avg_g2 = avg([g2_total.get(nm,0) for nm in l2_union_names])
            avg_g3 = avg([g3_total.get(nm,0) for nm in s3_union_names])
            avg_g4 = avg([g4_total.get(nm,0) for nm in group4_level_names])

            # Build one-column stats with grouped sections
            rows = []
            def add_group(section_name, members, getter, avg_val):
                for nm in members:
                    rows.append({"Surgeon": nm, "Group": section_name, "Number of past calls": int(getter(nm))})
                # average row
                rows.append({"Surgeon": f"Average ({section_name})", "Group": section_name, "Number of past calls": avg_val})
                # spacer
                rows.append({"Surgeon": "", "Group": "", "Number of past calls": ""})

            add_group("G1 (1A+1B)", group1_names, lambda nm: g1_total.get(nm,0), avg_g1)
            add_group("G2 (2A+2B)", l2_union_names, lambda nm: g2_total.get(nm,0), avg_g2)
            add_group("G3 (3 + 2B if grp4)", s3_union_names, lambda nm: g3_total.get(nm,0), avg_g3)
            add_group("G4 (4)", group4_level_names, lambda nm: g4_total.get(nm,0), avg_g4)

            import pandas as _pd
            df_hy_stats = _pd.DataFrame(rows)
        except Exception:
            df_hy_stats = None

        # 4) Build unavailability DataFrame
        availability = get_availability_requests()
        surgeons = get_all_surgeons()
        id_to_name = {s['id']: s['name'] for s in surgeons}
        data_unav = []
        for sid, reqs in availability.items():
            for req in reqs:
                if req['date'] in schedule:
                    data_unav.append({
                        'Surgeon': id_to_name.get(sid, ''),
                        'Date': req['date'],
                        'Type': req['request_type']
                    })
        df_unav = pd.DataFrame(data_unav)

        # 5) Gather public holidays
        hk_h = holidays.HK(years=[year])
        ph = {d.isoformat() for d in hk_h if d.year == year and d.month == month}

        # 6) Write to Excel with formatting
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df_sched.to_excel(writer, sheet_name='Call Schedule', index=False)
            df_stats.to_excel(writer, sheet_name='Monthly Stats', index=False)
            df_unav.to_excel(writer, sheet_name='Unavailability', index=False)

            wb = writer.book
            from openpyxl.styles import Font, PatternFill, Alignment
            # Style Call Schedule sheet
            ws1 = writer.sheets['Call Schedule']
            header_font = Font(bold=True, size=12)
            weekend_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            ph_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

            # Format header
            for cell in ws1[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            # Format rows
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=1):
                cell = row[0]
                date_str = cell.value
                try:
                    dt = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                except:
                    continue
                fill = weekend_fill if dt.weekday() >= 5 else (ph_fill if date_str in ph else None)
                if fill:
                    for c in ws1[cell.row]:
                        c.fill = fill
            # Adjust column widths
            for col in ws1.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                ws1.column_dimensions[col[0].column_letter].width = max_length + 2

            # Apply simple styling to other sheets (headers bold)
            for name in ['Monthly Stats', 'Unavailability']:
                ws = writer.sheets[name]
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

            # Append Half-Year grouped stats to the right of Monthly Stats table in one column
            if df_hy_stats is not None:
                ws_ms = writer.sheets['Monthly Stats']
                start_row = 1
                start_col = ws_ms.max_column + 2  # leave one blank column
                # Header for the single column
                ws_ms.cell(row=start_row, column=start_col, value='Number of past calls').font = header_font
                ws_ms.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
                # Group averages map for coloring
                avg_map = {
                    'G1 (1A+1B)': avg_g1 if 'avg_g1' in locals() else 0.0,
                    'G2 (2A+2B)': avg_g2 if 'avg_g2' in locals() else 0.0,
                    'G3 (3 + 2B if grp4)': avg_g3 if 'avg_g3' in locals() else 0.0,
                    'G4 (4)': avg_g4 if 'avg_g4' in locals() else 0.0,
                }
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                # Build row map by group to align with existing stats order
                # Create a mapping surgeon -> number of past calls and surgeon -> group label
                hy_map = {}
                group_map = {}
                for _, row in df_hy_stats.iterrows():
                    nm = row.get('Surgeon')
                    grp = row.get('Group')
                    if nm and not (isinstance(nm, str) and nm.startswith('Average (')):
                        hy_map[nm] = row.get('Number of past calls')
                        group_map[nm] = grp
                # Write aligned values next to existing table
                # Find Surgeon column in Monthly Stats
                hdr_to_col = {ws_ms.cell(row=1, column=c).value: c for c in range(1, ws_ms.max_column+1)}
                col_surgeon = hdr_to_col.get('Surgeon')
                r = 2
                while r <= ws_ms.max_row:
                    nm = ws_ms.cell(row=r, column=col_surgeon).value if col_surgeon else None
                    val = hy_map.get(nm)
                    ws_ms.cell(row=r, column=start_col, value=val if val is not None else "")
                    # Color row based on group average
                    try:
                        grp = group_map.get(nm)
                        if grp and isinstance(val, (int, float)):
                            avg_val = avg_map.get(grp, 0.0)
                            cell = ws_ms.cell(row=r, column=start_col)
                            if val < avg_val:
                                cell.fill = green_fill
                            elif val > avg_val:
                                cell.fill = red_fill
                    except Exception:
                        pass
                    r += 1
                # Insert an Average row after each group's last row
                group_last_row = {}
                for rr in range(2, ws_ms.max_row + 1):
                    nm_rr = ws_ms.cell(row=rr, column=col_surgeon).value if col_surgeon else None
                    grp_rr = group_map.get(nm_rr)
                    if grp_rr:
                        group_last_row[grp_rr] = rr
                groups_in_order = [
                    ('G1 (1A+1B)', avg_g1 if 'avg_g1' in locals() else 0.0),
                    ('G2 (2A+2B)', avg_g2 if 'avg_g2' in locals() else 0.0),
                    ('G3 (3 + 2B if grp4)', avg_g3 if 'avg_g3' in locals() else 0.0),
                    ('G4 (4)', avg_g4 if 'avg_g4' in locals() else 0.0),
                ]
                to_insert = [(group_last_row[g], g, a) for g, a in groups_in_order if g in group_last_row]
                to_insert.sort(key=lambda x: x[0], reverse=True)
                for base_row, grp_label, avg_val in to_insert:
                    insert_at = base_row + 1
                    ws_ms.insert_rows(insert_at, amount=1)
                    c_lbl = ws_ms.cell(row=insert_at, column=col_surgeon or 1, value=f"Average ({grp_label})")
                    c_lbl.font = header_font
                    ws_ms.cell(row=insert_at, column=start_col, value=avg_val)
                # Autosize the appended single column width
                col_cells = [ws_ms.cell(row=i, column=start_col) for i in range(1, ws_ms.max_row + 1)]
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws_ms.column_dimensions[col_cells[0].column_letter].width = max_len + 2

        bio.seek(0)
        filename = f"call_list_{year}_{month:02d}.xlsx"
        return send_file(
            bio,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )


    @app.route('/')
    def index():
        year_sel, month_sel = get_year_month()
        db = get_db()
        row = db.execute(
            text("SELECT * FROM saved_schedule WHERE year = :year AND month = :month"),
            {"year": year_sel, "month": month_sel}
        ).fetchone()
        if row:
            return redirect(url_for('saved_schedule', year=year_sel, month=month_sel))
        else:
            return redirect(url_for('saved_schedule', year=year_sel, month=month_sel))

    #############################################
    # Availability / Unavailability Endpoint
    #############################################

    @app.route('/availability', methods=['GET', 'POST'])
    def availability():
        db = get_db()
        now = datetime.datetime.now()
        deadline_dt, deadline_passed = is_unavail_deadline_passed(now=now)
        next_month_year = now.year + (1 if now.month == 12 else 0)
        next_month_month = 1 if now.month == 12 else now.month + 1
        next_month_label = datetime.date(next_month_year, next_month_month, 1).strftime("%B %Y")

        def is_admin_request():
            auth = request.authorization
            expected_user = app.config.get('BASIC_AUTH_USERNAME')
            expected_pass = app.config.get('BASIC_AUTH_PASSWORD')
            return bool(auth and expected_user and expected_pass and auth.username == expected_user and auth.password == expected_pass)

        is_admin = is_admin_request()
        block_requests = bool(deadline_dt and deadline_passed and not is_admin)
        next_month_start = datetime.date(next_month_year, next_month_month, 1)
        next_month_end = (next_month_start + datetime.timedelta(days=32)).replace(day=1)
        if request.method == 'POST':
            # Process submission of selected dates.
            surgeon_id = request.form.get('surgeon_id')
            request_type = request.form.get('request_type')  # "unavailable" or "no_call"
            selected_dates_json = request.form.get('selected_dates')

            # Validate that all required data is provided.
            if not surgeon_id or not request_type or not selected_dates_json:
                flash("Please fill in all fields.")
                return redirect(url_for('availability', surgeon_id=surgeon_id))
            
            try:
                # Parse the JSON array of dates.
                raw_dates = json.loads(selected_dates_json)
                dates_list = raw_dates if isinstance(raw_dates, list) else []
                if not dates_list:
                    flash("No dates selected.")
                    return redirect(url_for('availability', surgeon_id=surgeon_id))
                def _coerce_date(val):
                    if isinstance(val, datetime.date):
                        return val
                    return datetime.date.fromisoformat(val)
                if block_requests:
                    for d in dates_list:
                        try:
                            d_obj = _coerce_date(d)
                        except Exception:
                            raise
                        if next_month_start <= d_obj < next_month_end:
                            flash(f"Requests for {next_month_label} are closed after the deadline. Later months remain open.", "warning")
                            return redirect(url_for('availability', surgeon_id=surgeon_id))
                
                # Insert each date as a separate row using a fresh transaction to avoid conflicts.
                with ENGINE.begin() as write_conn:
                    for d in dates_list:
                        write_conn.execute(
                            text(
                                "INSERT INTO surgeon_availability "
                                "(surgeon_id, request_type, date) "
                                "VALUES (:sid, :rtype, :dt)"
                            ),
                            {"sid": surgeon_id, "rtype": request_type, "dt": d}
                        )
                flash("Request submitted successfully!")
            except Exception as e:
                flash(f"Error processing the dates: {str(e)}")
            return redirect(url_for('availability', surgeon_id=surgeon_id))
        else:
            # GET: Display the page with existing requests grouped by date range.
            surgeon_id = request.args.get('surgeon_id')
            try:
                surgeon_id = int(surgeon_id) if surgeon_id else None
            except ValueError:
                surgeon_id = None

            events = {}
            surgeon_name = ""
            if surgeon_id:
                rows = (
                    db.execute(
                        text(
                            "SELECT sa.date, sa.request_type, s.name, s.call_levels "
                            "FROM surgeon_availability sa "
                            "JOIN surgeons s ON sa.surgeon_id = s.id "
                            "WHERE s.id = :sid "
                            "ORDER BY sa.date"
                        ),
                        {"sid": surgeon_id}
                    )
                    .mappings()
                    .all()
                )
                if rows:
                    surgeon_name = rows[0]["name"]
                grouped_by_type = {}
                for row in rows:
                    rtype = row["request_type"]
                    if rtype not in grouped_by_type:
                        grouped_by_type[rtype] = []
                    grouped_by_type[rtype].append(row["date"])
                for rtype, date_list in grouped_by_type.items():
                    events[rtype] = group_dates(date_list)
        surgeons = get_all_surgeons()
        return render_template('availability.html', events=events, surgeons=surgeons,
                               selected_surgeon_id=surgeon_id, surgeon_name=surgeon_name,
                               deadline_blocked=block_requests,
                               deadline_dt=deadline_dt,
                               deadline_passed=bool(deadline_dt and deadline_passed),
                               deadline_next_month_label=next_month_label,
                               is_admin=is_admin)

    #############################################
    # Delete Availability Request Endpoint
    #############################################

    @app.route('/delete_availability', methods=['POST'])
    def delete_availability():
        surgeon_id = request.form.get('surgeon_id')
        delete_requests_json = request.form.get('delete_requests')
        
        if not surgeon_id or not delete_requests_json:
            flash("Missing surgeon id or delete requests data.", category="error")
            return redirect(url_for('availability', surgeon_id=surgeon_id))
        
        try:
            delete_requests = json.loads(delete_requests_json)
        except Exception:
            flash("Invalid delete requests data.", category="error")
            return redirect(url_for('availability', surgeon_id=surgeon_id))
        
        db = get_db()  # Get the DB connection from your helper

        with ENGINE.begin() as write_conn:
            success = True
            for req in delete_requests:
                req_type = req.get('reqType')
                start_date = req.get('start')
                end_date = req.get('end')
                if not (req_type and start_date and end_date):
                    continue
                # Delete all rows within the given date range.
                query = text("""
                    DELETE FROM surgeon_availability 
                    WHERE surgeon_id = :surgeon_id 
                    AND request_type = :req_type 
                    AND date BETWEEN :start_date AND :end_date
                """)
                result = write_conn.execute(query, {
                    "surgeon_id": surgeon_id,
                    "req_type": req_type,
                    "start_date": start_date,
                    "end_date": end_date
                })
                if result.rowcount == 0:
                    success = False
            
            if success:
                flash("Selected availability requests deleted successfully.", category="success")
            else:
                flash("Some availability requests could not be deleted.", category="warning")
            
        return redirect(url_for('availability', surgeon_id=surgeon_id))
    
    #############################################
    # Export availability Endpoint
    #############################################

    @app.route('/export_requests', methods=['POST'])
    @basic_auth.required
    def export_requests():
        try:
            year = int(request.form.get("year"))
            month = int(request.form.get("month"))
        except (TypeError, ValueError):
            flash("Invalid year or month provided.", category="error")
            return redirect(url_for('new_schedule'))

        import calendar as _cal
        from datetime import date
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        num_days = _cal.monthrange(year, month)[1]
        days = [date(year, month, d) for d in range(1, num_days + 1)]

        hk_h = holidays.HK(years=[year])
        holiday_dates = {d for d in hk_h if d.year == year and d.month == month}
        weekend_idx = {d.day for d in days if d.weekday() >= 5}

        all_surgeons = get_all_surgeons()
        level_rank = {"1A": 1, "1B": 1, "2A": 2, "2B": 3, "3": 4, "4": 5}

        def get_call_rank(call_levels):
            levels = parse_call_levels(call_levels or "")
            if not levels:
                return 99
            return min(level_rank.get(l, 99) for l in levels)

        def call_levels_label(call_levels):
            levels = parse_call_levels(call_levels or "")
            return ",".join(sorted(levels, key=lambda l: level_rank.get(l, 99))) if levels else ""

        def get_primary_level(call_levels):
            levels = parse_call_levels(call_levels or "")
            if not levels:
                return "Other"
            best = min(levels, key=lambda l: level_rank.get(l, 99))
            return best

        start_date = date(year, month, 1)
        next_date = date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1)
        db = get_db()
        stmt = text(
            """
            SELECT surgeon_id, request_type, date
            FROM surgeon_availability
            WHERE date >= :start_date AND date < :next_date
            """
        )
        req_rows = db.execute(stmt, {"start_date": start_date.isoformat(), "next_date": next_date.isoformat()}).mappings().all()

        marks = {}
        for r in req_rows:
            rtype = r['request_type']
            if rtype in ('unavailable', 'no_call', 'study_leave'):
                marks[(r['surgeon_id'], r['date'].day)] = rtype

        # Shared styles
        dark_red = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        weekend_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        holiday_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        pct_green = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        pct_yellow = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        pct_red = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        header_font = Font(bold=True)
        summary_font = Font(bold=True, size=10)
        white_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin_border_top = Border(top=Side(style='thin'))

        def _write_header(ws):
            ws.cell(row=1, column=1, value="Surgeon")
            ws.cell(row=1, column=2, value="Levels")
            for d in range(1, num_days + 1):
                ws.cell(row=1, column=2 + d, value=d)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center

        def _apply_weekend_holiday_bg(ws, last_row):
            for d in range(1, num_days + 1):
                col_idx = 2 + d
                the_date = date(year, month, d)
                base_fill = None
                if the_date in holiday_dates:
                    base_fill = holiday_fill
                elif d in weekend_idx:
                    base_fill = weekend_fill
                if base_fill:
                    for r in range(2, last_row + 1):
                        c = ws.cell(row=r, column=col_idx)
                        if c.fill == PatternFill():
                            c.fill = base_fill

        def _fill_marks(ws, surgeon_id_to_row):
            for (sid, d), rtype in marks.items():
                row = surgeon_id_to_row.get(sid)
                if not row:
                    continue
                cell = ws.cell(row=row, column=2 + d)
                if rtype == 'unavailable':
                    cell.value = 'U'
                    cell.fill = dark_red
                    cell.font = white_font
                elif rtype == 'no_call':
                    cell.value = 'N'
                    cell.fill = orange_fill
                elif rtype == 'study_leave':
                    cell.value = 'SL'
                    cell.fill = green_fill
                cell.alignment = center

        def _autosize(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(18, max(6, max_len + 2))

        # ---- Tab 1: By Team → Level ----
        surgeons_team = sorted(all_surgeons, key=lambda s: (
            (s.get('team') or 'zzz'),
            get_call_rank(s.get('call_levels')),
            s.get('name', '').lower()
        ))
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "By Team"
        _write_header(ws1)

        id_to_row_t = {}
        row_cursor = 2
        prev_team = None
        for s in surgeons_team:
            team = s.get('team')
            if prev_team is not None and team != prev_team:
                row_cursor += 1
            id_to_row_t[s['id']] = row_cursor
            ws1.cell(row=row_cursor, column=1, value=f"{team or ''} - {s.get('name')}")
            ws1.cell(row=row_cursor, column=2, value=call_levels_label(s.get('call_levels')))
            row_cursor += 1
            prev_team = team
        last_row_t = row_cursor - 1

        _fill_marks(ws1, id_to_row_t)
        _apply_weekend_holiday_bg(ws1, last_row_t)
        _autosize(ws1)

        # ---- Tab 2: By Level → Team (with availability % summary rows) ----
        surgeons_level = sorted(all_surgeons, key=lambda s: (
            get_call_rank(s.get('call_levels')),
            (s.get('team') or 'zzz'),
            s.get('name', '').lower()
        ))
        ws2 = wb.create_sheet("By Level")
        _write_header(ws2)

        id_to_row_l = {}
        summary_rows = []
        row_cursor = 2
        prev_rank = None
        group_start = 2
        group_ids = []

        for s in surgeons_level:
            rank = get_call_rank(s.get('call_levels'))
            if prev_rank is not None and rank != prev_rank:
                summary_rows.append((row_cursor, list(group_ids), prev_rank))
                row_cursor += 1
                row_cursor += 1
                group_ids = []
                group_start = row_cursor
            id_to_row_l[s['id']] = row_cursor
            ws2.cell(row=row_cursor, column=1, value=f"{s.get('name')} ({s.get('team') or ''})")
            ws2.cell(row=row_cursor, column=2, value=call_levels_label(s.get('call_levels')))
            group_ids.append(s['id'])
            row_cursor += 1
            prev_rank = rank

        if group_ids:
            summary_rows.append((row_cursor, list(group_ids), prev_rank))
            row_cursor += 1

        last_row_l = row_cursor - 1

        _fill_marks(ws2, id_to_row_l)

        rank_to_label = {}
        for lvl, rk in level_rank.items():
            rank_to_label.setdefault(rk, []).append(lvl)
        for rk in rank_to_label:
            rank_to_label[rk] = "/".join(sorted(rank_to_label[rk]))

        for summary_row, sids, rank_val in summary_rows:
            pool_size = len(sids)
            label = rank_to_label.get(rank_val, "Other")
            ws2.cell(row=summary_row, column=1, value=f"% Available ({label})")
            ws2.cell(row=summary_row, column=1).font = summary_font
            ws2.cell(row=summary_row, column=1).border = thin_border_top

            for d in range(1, num_days + 1):
                unavail_count = sum(1 for sid in sids if (sid, d) in marks)
                avail_pct = ((pool_size - unavail_count) / pool_size * 100) if pool_size > 0 else 0
                cell = ws2.cell(row=summary_row, column=2 + d)
                cell.value = f"{avail_pct:.0f}%"
                cell.alignment = center
                cell.font = summary_font
                cell.border = thin_border_top
                if avail_pct > 50:
                    cell.fill = pct_green
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                elif avail_pct > 25:
                    cell.fill = pct_yellow
                else:
                    cell.fill = pct_red
                    cell.font = Font(bold=True, color="FFFFFF", size=10)

        _apply_weekend_holiday_bg(ws2, last_row_l)
        _autosize(ws2)

        # ---- Tab 3: By Name (A-Z) ----
        surgeons_name = sorted(all_surgeons, key=lambda s: s.get('name', '').lower())
        ws3 = wb.create_sheet("By Name")
        _write_header(ws3)

        id_to_row_n = {}
        row_cursor = 2
        for s in surgeons_name:
            id_to_row_n[s['id']] = row_cursor
            ws3.cell(row=row_cursor, column=1, value=s.get('name'))
            ws3.cell(row=row_cursor, column=2, value=call_levels_label(s.get('call_levels')))
            row_cursor += 1
        last_row_n = row_cursor - 1

        _fill_marks(ws3, id_to_row_n)
        _apply_weekend_holiday_bg(ws3, last_row_n)
        _autosize(ws3)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        filename = f"requests_matrix_{year}_{month:02d}.xlsx"
        return send_file(
            out,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route('/eligible_for_day')
    @basic_auth.required
    def eligible_for_day():
        # Return per-level eligible candidate lists for the given date based on availability/no_call only
        day = request.args.get('day')
        try:
            target = datetime.date.fromisoformat(day)
        except Exception:
            return jsonify({"error": "Invalid day"}), 400
        surgeons = get_all_surgeons()
        availability = get_availability_requests()
        # parse levels
        def has_level(s, lvl):
            return lvl in parse_call_levels(s.get('call_levels',''))
        result = {lvl: [] for lvl in ["1A","1B","2A","2B","3","4"]}
        for lvl in result.keys():
            for s in surgeons:
                if not has_level(s, lvl):
                    continue
                sid = s['id']
                # exclude if unavailable/no_call exactly on target day
                bad = False
                for req in availability.get(sid, []):
                    raw = req.get('date')
                    d = raw if isinstance(raw, datetime.date) else None
                    if not d:
                        try:
                            d = datetime.date.fromisoformat(raw)
                        except Exception:
                            continue
                if d == target and req.get('request_type') in ('unavailable','study_leave','no_call'):
                        bad = True
                        break
                if not bad:
                    result[lvl].append({"id": sid, "name": s['name']})
        return jsonify(result)

    @app.route('/eligible_for_month')
    @basic_auth.required
    def eligible_for_month():
        try:
            year = int(request.args.get('year'))
            month = int(request.args.get('month'))
        except Exception:
            return jsonify({"error": "Invalid parameters"}), 400
        import calendar as _cal
        num_days = _cal.monthrange(year, month)[1]
        days = [datetime.date(year, month, d) for d in range(1, num_days + 1)]
        surgeons = get_all_surgeons()
        availability = get_availability_requests()
        # Precompute blocked days per surgeon (unavailable or no_call in this month)
        blocked = {}
        for s in surgeons:
            sid = s['id']
            bset = set()
            for req in availability.get(sid, []):
                raw = req.get('date')
                d = raw if isinstance(raw, datetime.date) else None
                if not d:
                    try:
                        d = datetime.date.fromisoformat(raw)
                    except Exception:
                        continue
                if d.year == year and d.month == month and req.get('request_type') in ('unavailable', 'study_leave', 'no_call'):
                    bset.add(d)
            blocked[sid] = bset
        # Precompute level eligibility per surgeon
        def has_level(s, lvl):
            return lvl in parse_call_levels(s.get('call_levels',''))
        levels = ["1A","1B","2A","2B","3","4"]
        base_by_level = {lvl: [s for s in surgeons if has_level(s, lvl)] for lvl in levels}
        # Build response map
        res = {}
        for d in days:
            dkey = d.isoformat()
            res[dkey] = {}
            for lvl in levels:
                elig = []
                for s in base_by_level[lvl]:
                    if d not in blocked.get(s['id'], set()):
                        elig.append({"id": s['id'], "name": s['name']})
                res[dkey][lvl] = elig
        return jsonify(res)

    # Provide surgeon id/name lookup for UI reconciliation
    @app.route('/surgeons_lookup')
    @basic_auth.required
    def surgeons_lookup():
        surgeons = get_all_surgeons()
        by_id = {int(s['id']): s['name'] for s in surgeons}
        by_name = {s['name']: int(s['id']) for s in surgeons}
        return jsonify({"by_id": by_id, "by_name": by_name})

    @app.route('/edit_month_data')
    @basic_auth.required
    def edit_month_data():
        """Comprehensive data for the edit/publish page: eligibility, availability details, surgeon list, prior-month carry-over."""
        try:
            year = int(request.args.get('year'))
            month = int(request.args.get('month'))
        except Exception:
            return jsonify({"error": "Invalid parameters"}), 400

        import calendar as _cal
        num_days = _cal.monthrange(year, month)[1]
        days = [datetime.date(year, month, d) for d in range(1, num_days + 1)]
        surgeons = get_all_surgeons()
        availability = get_availability_requests()

        blocked = {}
        blocked_reasons = {}
        for s in surgeons:
            sid = s['id']
            bset = set()
            rmap = {}
            for req in availability.get(sid, []):
                raw = req.get('date')
                d = raw if isinstance(raw, datetime.date) else None
                if not d:
                    try:
                        d = datetime.date.fromisoformat(raw)
                    except Exception:
                        continue
                if d.year == year and d.month == month:
                    rtype = req.get('request_type', '')
                    if rtype in ('unavailable', 'study_leave', 'no_call'):
                        bset.add(d)
                        rmap[d.isoformat()] = rtype
            blocked[sid] = bset
            blocked_reasons[sid] = rmap

        def has_level(s, lvl):
            return lvl in parse_call_levels(s.get('call_levels', ''))

        levels = ["1A", "1B", "2A", "2B", "3", "4"]
        base_by_level = {lvl: [s for s in surgeons if has_level(s, lvl)] for lvl in levels}

        eligibility = {}
        avail_map = {}
        for d in days:
            dkey = d.isoformat()
            eligibility[dkey] = {}
            day_avail = {}
            for s in surgeons:
                sid = s['id']
                reason = blocked_reasons.get(sid, {}).get(dkey)
                day_avail[str(sid)] = reason if reason else "available"
            avail_map[dkey] = day_avail
            for lvl in levels:
                elig = []
                for s in base_by_level[lvl]:
                    if d not in blocked.get(s['id'], set()):
                        elig.append({"id": s['id'], "name": s['name']})
                eligibility[dkey][lvl] = elig

        all_surgeons_list = [
            {"id": s['id'], "name": s['name'],
             "call_levels": s.get('call_levels', ''),
             "team": s.get('team', '')}
            for s in surgeons
        ]

        prior = get_prior_last_two(year, month)

        hk_h = holidays.HK(years=[year])
        public_holidays_list = sorted(
            d.isoformat() for d in hk_h
            if d.year == year and d.month == month
        )

        return jsonify({
            "eligibility": eligibility,
            "availability": avail_map,
            "all_surgeons": all_surgeons_list,
            "prior_last_two": prior,
            "public_holidays": public_holidays_list
        })

    #############################################
    # Delete Surgeon Endpoint
    #############################################

    @app.route('/surgeons/delete/<int:surgeon_id>', methods=['POST'])
    @basic_auth.required
    def delete_surgeon(surgeon_id):
        db = get_db()
        with db.begin():
        # First, clean up any related availability records:
            db.execute(
                text("DELETE FROM surgeon_availability WHERE surgeon_id = :sid"),
                {"sid": surgeon_id}
            )
            # Then delete the surgeon:
            db.execute(
                text("DELETE FROM surgeons WHERE id = :sid"),
                {"sid": surgeon_id}
            )
            flash("Surgeon deleted successfully.", "success")
            return redirect(url_for('list_surgeons'))


    #############################################
    # Stats endpoint
    #############################################

    @app.route('/stats', methods=['GET'])
    @basic_auth.required
    def stats():
        import datetime
        # Get the start and end period from query parameters.
        try:
            start_year = int(request.args.get('start_year', datetime.date.today().year))
            start_month = int(request.args.get('start_month', datetime.date.today().month))
            end_year = int(request.args.get('end_year', datetime.date.today().year))
            end_month = int(request.args.get('end_month', datetime.date.today().month))
        except ValueError:
            flash("Invalid date range provided.")
            return redirect(url_for('stats'))
        
        # We assume that each saved_schedule record has integer columns "year" and "month".
        # Query the saved_schedule table for records in the selected range.
        db = get_db()
        # One simple way: only include records where (year, month) is within the selected range.
        # Here we compare first by year, then by month.
        stmt = text("""
            SELECT year, month, schedule_data
            FROM saved_schedule
            WHERE (year > :sy OR (year = :sy AND month >= :sm))
            AND (year < :ey OR (year = :ey AND month <= :em))
            ORDER BY year, month
        """)
        result  = db.execute(stmt, {
            "sy": start_year, "sm": start_month,
            "ey": end_year,   "em": end_month
        })
        records = result.mappings().all()
        
        # Define mapping for call level ranking.
        level_ranks = {
            "1A": 1,
            "1B": 1,
            "2A": 2,
            "2B": 3,
            "3": 4,
            "4": 5
        }
        
        # Initialize dictionary to aggregate stats for each surgeon.
        # We use surgeon names as keys.
        stats_dict = {}
        
        for row in records:
            raw = row["schedule_data"]
            # raw might already be a dict (Postgres JSONB) or a JSON string
            try:
                sched = raw if isinstance(raw, dict) else json.loads(raw)
            except Exception:
                # skip corrupted entries
                continue

            for day_str, assigns in sched.items():
                try:
                    d = datetime.date.fromisoformat(day_str)
                except Exception:
                    continue
                is_weekend = (d.weekday() >= 5)
                for lvl, surgeon in assigns.items():
                    if not surgeon:
                        continue
                    rec = stats_dict.setdefault(surgeon, {
                        "total_calls": 0,
                        "weekend_calls": 0,
                        "min_level_rank": None
                    })
                    rec["total_calls"] += 1
                    if is_weekend:
                        rec["weekend_calls"] += 1
                    rank = level_ranks.get(lvl, 99)
                    if rec["min_level_rank"] is None or rank < rec["min_level_rank"]:
                        rec["min_level_rank"] = rank

        # Turn into a sorted list for the template
        stats_list = sorted(
            [
                {
                    "surgeon": name,
                    "total_calls": data["total_calls"],
                    "weekend_calls": data["weekend_calls"],
                    "min_level_rank": data["min_level_rank"]
                }
                for name, data in stats_dict.items()
            ],
            key=lambda x: (x["min_level_rank"], x["surgeon"].lower())
        )

        return render_template(
            'stats.html',
            stats=stats_list,
            start_year=start_year,
            start_month=start_month,
            end_year=end_year,
            end_month=end_month
        )

    #############################################
    # Edit and Publish (Versioned schedules)
    #############################################

    @app.route('/edit_publish', methods=['GET'])
    @basic_auth.required
    def edit_publish():
        try:
            # Default to next month when no params supplied
            if not request.args.get('year') or not request.args.get('month'):
                today = datetime.date.today()
                year_sel = today.year + 1 if today.month == 12 else today.year
                month_sel = 1 if today.month == 12 else today.month + 1
            else:
                year_sel, month_sel = get_year_month()
            db = get_db()
            versions = db.execute(text("SELECT version, version_name, published FROM saved_schedule_versions WHERE year=:y AND month=:m ORDER BY version"), {"y": year_sel, "m": month_sel}).mappings().all()
            if not versions:
                # Seed v1 from existing saved_schedule if present; otherwise create empty skeleton
                row = db.execute(text("SELECT schedule_data FROM saved_schedule WHERE year=:y AND month=:m"), {"y": year_sel, "m": month_sel}).mappings().fetchone()
                data = coerce_json_column(row['schedule_data']) if row else None
                if data is None:
                    # build empty schedule for the month
                    import calendar as _cal
                    num_days = _cal.monthrange(year_sel, month_sel)[1]
                    levels = ["1A","1B","2A","2B","3","4"]
                    data = { datetime.date(year_sel, month_sel, d).isoformat(): {lvl: None for lvl in levels} for d in range(1, num_days+1) }
                
                # Commit any implicit transaction from previous reads before starting the write block
                db.commit()
                # Use a new connection for the transaction to avoid state issues
                with db.begin():
                    # Attempt to handle potential dialect issues with JSONB
                    db.execute(
                        text(
                            f"INSERT INTO saved_schedule_versions (year, month, version, version_name, schedule_data, published) "
                            f"VALUES (:y, :m, 1, :vn, {json_cast('d')}, {sql_false()})"
                        ),
                        {"y": year_sel, "m": month_sel, "vn": "v1", "d": json.dumps(data)}
                    )
                versions = db.execute(text("SELECT version, version_name, published FROM saved_schedule_versions WHERE year=:y AND month=:m ORDER BY version"), {"y": year_sel, "m": month_sel}).mappings().all()
            # latest published
            pub_ver = next((v['version'] for v in versions if v['published']), None)
            return render_template('edit_publish.html', year=year_sel, month=month_sel, versions=versions, published_version=pub_ver)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"Error in edit_publish: {e}", 500

    @app.route('/list_schedule_versions')
    @basic_auth.required
    def list_schedule_versions():
        try:
            year = int(request.args.get('year'))
            month = int(request.args.get('month'))
        except Exception:
            return jsonify({"error": "Invalid parameters"}), 400
        db = get_db()
        rows = db.execute(text("SELECT version, version_name, published FROM saved_schedule_versions WHERE year=:y AND month=:m ORDER BY version"), {"y": year, "m": month}).mappings().all()
        if not rows:
            # Seed v1 like the page handler does
            row = db.execute(text("SELECT schedule_data FROM saved_schedule WHERE year=:y AND month=:m"), {"y": year, "m": month}).mappings().fetchone()
            data = coerce_json_column(row['schedule_data']) if row else None
            if data is None:
                import calendar as _cal
                num_days = _cal.monthrange(year, month)[1]
                levels = ["1A","1B","2A","2B","3","4"]
                data = { datetime.date(year, month, d).isoformat(): {lvl: None for lvl in levels} for d in range(1, num_days+1) }
            
            db.commit()
            with db.begin():
                db.execute(
                    text(
                        f"INSERT INTO saved_schedule_versions (year, month, version, version_name, schedule_data, published) "
                        f"VALUES (:y, :m, 1, :vn, {json_cast('d')}, {sql_false()})"
                    ),
                    {"y": year, "m": month, "vn": "v1", "d": json.dumps(data)}
                )
            rows = db.execute(text("SELECT version, version_name, published FROM saved_schedule_versions WHERE year=:y AND month=:m ORDER BY version"), {"y": year, "m": month}).mappings().all()
        return jsonify({
            "versions": [
                {
                    "version": r['version'],
                    "name": (r.get('version_name') or f"v{r['version']}"),
                    "published": bool(r['published'])
                }
                for r in rows
            ]
        })

    @app.route('/rename_schedule_version', methods=['POST'])
    @basic_auth.required
    def rename_schedule_version():
        data = request.get_json() or {}
        try:
            year = int(data.get('year'))
            month = int(data.get('month'))
            version = int(data.get('version'))
            version_name = str(data.get('version_name') or '').strip()
        except Exception:
            return jsonify({"error": "Invalid payload"}), 400
        if not version_name:
            return jsonify({"error": "Version name cannot be empty"}), 400
        db = get_db()
        with db.begin():
            res = db.execute(
                text(
                    f"UPDATE saved_schedule_versions "
                    f"SET version_name = :vn, updated_at = {sql_now()} "
                    f"WHERE year=:y AND month=:m AND version=:v"
                ),
                {"vn": version_name, "y": year, "m": month, "v": version}
            )
        if res.rowcount == 0:
            return jsonify({"error": "Version not found"}), 404
        return jsonify({"ok": True, "version": version, "version_name": version_name})

    @app.route('/publish_schedule_version', methods=['POST'])
    @basic_auth.required
    def publish_schedule_version():
        data = request.get_json() or {}
        try:
            year = int(data.get('year'))
            month = int(data.get('month'))
            version = int(data.get('version'))
            published = bool(data.get('published', True))
        except Exception:
            return jsonify({"error": "Invalid payload"}), 400
        db = get_db()
        with db.begin():
            if published:
                db.execute(text(f"UPDATE saved_schedule_versions SET published = {sql_false()} WHERE year=:y AND month=:m"), {"y": year, "m": month})
                res = db.execute(text(f"UPDATE saved_schedule_versions SET published = {sql_true()} WHERE year=:y AND month=:m AND version=:v"), {"y": year, "m": month, "v": version})
            else:
                res = db.execute(text(f"UPDATE saved_schedule_versions SET published = {sql_false()} WHERE year=:y AND month=:m AND version=:v"), {"y": year, "m": month, "v": version})
        if res.rowcount == 0:
            return jsonify({"error": "Version not found"}), 404
        return jsonify({"ok": True})

    @app.route('/delete_schedule_version', methods=['POST'])
    @basic_auth.required
    def delete_schedule_version():
        data = request.get_json() or {}
        try:
            year = int(data.get('year'))
            month = int(data.get('month'))
            version = int(data.get('version'))
        except Exception:
            return jsonify({"error": "Invalid payload"}), 400
        db = get_db()
        with db.begin():
            res = db.execute(text("DELETE FROM saved_schedule_versions WHERE year=:y AND month=:m AND version=:v"), {"y": year, "m": month, "v": version})
        if res.rowcount == 0:
            return jsonify({"error": "Version not found"}), 404
        return jsonify({"ok": True})

    @app.route('/load_schedule_version')
    @basic_auth.required
    def load_schedule_version():
        try:
            year = int(request.args.get('year'))
            month = int(request.args.get('month'))
            version = int(request.args.get('version'))
        except Exception:
            return jsonify({"error": "Invalid parameters"}), 400
        db = get_db()
        row = db.execute(text("SELECT schedule_data FROM saved_schedule_versions WHERE year=:y AND month=:m AND version=:v"), {"y": year, "m": month, "v": version}).mappings().fetchone()
        if not row:
            return jsonify({"error": "Not found"}), 404
        data = coerce_json_column(row['schedule_data'])
        if data is None:
            return jsonify({"error": "Invalid or empty schedule data"}), 400
        return jsonify({"schedule": data})

    @app.route('/preassignments_for_month')
    @basic_auth.required
    def preassignments_for_month():
        try:
            year = int(request.args.get('year'))
            month = int(request.args.get('month'))
        except Exception:
            return jsonify({"error": "Invalid parameters"}), 400
        db = get_db()
        # Load raw preassignments JSON for the month
        row = db.execute(text("SELECT preassignment_data FROM preassignments WHERE year = :year AND month = :month"), {"year": year, "month": month}).mappings().fetchone()
        if not row or not row['preassignment_data']:
            return jsonify({})
        try:
            raw = row['preassignment_data']
            pre = raw if isinstance(raw, dict) else json.loads(raw)
        except Exception:
            return jsonify({})
        # Map surgeon IDs to names for convenience in the UI
        surgeons = get_all_surgeons()
        id_to_name = {s['id']: s['name'] for s in surgeons}
        enriched = {}
        for day, lvls in (pre or {}).items():
            if not isinstance(lvls, dict):
                continue
            enriched.setdefault(day, {})
            for lvl, sid in lvls.items():
                try:
                    si = int(sid) if sid not in [None, ""] else None
                except Exception:
                    si = None
                if si:
                    enriched[day][lvl] = {"id": si, "name": id_to_name.get(si, str(si))}
        return jsonify(enriched)

    @app.route('/save_schedule_version', methods=['POST'])
    @basic_auth.required
    def save_schedule_version():
        data = request.get_json()
        try:
            year = int(data.get('year'))
            month = int(data.get('month'))
            schedule = data.get('schedule') or {}
            publish = bool(data.get('publish', False))
            overwrite_version_raw = data.get('overwrite_version')
            overwrite_version = int(overwrite_version_raw) if overwrite_version_raw not in [None, ""] else None
            version_name = str(data.get('version_name') or '').strip()
        except Exception:
            return jsonify({"error": "Invalid payload"}), 400
        db = get_db()
        next_ver = 0
        final_name = ""
        with db.begin():
            rows = db.execute(text("SELECT id, name FROM surgeons")).mappings().all()
            id_to_name = {row["id"]: row["name"] for row in rows}
            sched_by_name = convert_schedule_ids_to_names(schedule, id_to_name)
            if overwrite_version is not None:
                exists = db.execute(
                    text("SELECT version, version_name FROM saved_schedule_versions WHERE year=:y AND month=:m AND version=:v"),
                    {"y": year, "m": month, "v": overwrite_version}
                ).mappings().fetchone()
                if not exists:
                    return jsonify({"error": "Overwrite target version not found"}), 404
                final_name = version_name or exists.get("version_name") or f"v{overwrite_version}"
                db.execute(
                    text(
                        f"UPDATE saved_schedule_versions "
                        f"SET version_name=:vn, schedule_data={json_cast('d')}, published=:p, updated_at={sql_now()} "
                        f"WHERE year=:y AND month=:m AND version=:v"
                    ),
                    {"vn": final_name, "d": json.dumps(sched_by_name), "p": publish, "y": year, "m": month, "v": overwrite_version}
                )
                next_ver = overwrite_version
            else:
                row = db.execute(text("SELECT COALESCE(MAX(version),0) AS maxv FROM saved_schedule_versions WHERE year=:y AND month=:m"),
                                 {"y": year, "m": month}).mappings().fetchone()
                next_ver = (row['maxv'] or 0) + 1
                final_name = version_name or f"v{next_ver}"
                db.execute(
                    text(f"INSERT INTO saved_schedule_versions (year, month, version, version_name, schedule_data, published) "
                         f"VALUES (:y, :m, :v, :vn, {json_cast('d')}, :p)"),
                    {"y": year, "m": month, "v": next_ver, "vn": final_name, "d": json.dumps(sched_by_name), "p": publish}
                )
            if publish:
                db.execute(
                    text(f"UPDATE saved_schedule_versions SET published = {sql_false()} "
                         "WHERE year=:y AND month=:m AND version <> :v"),
                    {"y": year, "m": month, "v": next_ver}
                )
        return jsonify({"version": next_ver, "version_name": final_name, "published": publish, "overwrote": overwrite_version is not None})

    @app.route('/save_schedule_version_form', methods=['POST'])
    @basic_auth.required
    def save_schedule_version_form():
        try:
            year = int(request.form.get('year'))
            month = int(request.form.get('month'))
            publish = request.form.get('publish') in ['1', 'true', 'True', 'on']
            sched_json = request.form.get('schedule_json_ids') or '{}'
            schedule = json.loads(sched_json)
            overwrite_version_raw = request.form.get('overwrite_version')
            overwrite_version = int(overwrite_version_raw) if overwrite_version_raw not in [None, ""] else None
            version_name = str(request.form.get('version_name') or '').strip()
        except Exception:
            flash('Invalid form submission for saving schedule version.', 'error')
            return redirect(url_for('edit_publish', year=request.form.get('year'), month=request.form.get('month')))

        db = get_db()
        next_ver = 0
        final_name = ""
        with db.begin():
            rows = db.execute(text("SELECT id, name FROM surgeons")).mappings().all()
            id_to_name = {row["id"]: row["name"] for row in rows}
            sched_by_name = convert_schedule_ids_to_names(schedule, id_to_name)
            if overwrite_version is not None:
                exists = db.execute(
                    text("SELECT version, version_name FROM saved_schedule_versions WHERE year=:y AND month=:m AND version=:v"),
                    {"y": year, "m": month, "v": overwrite_version}
                ).mappings().fetchone()
                if not exists:
                    flash('Overwrite target version not found.', 'error')
                    return redirect(url_for('edit_publish', year=year, month=month))
                final_name = version_name or exists.get("version_name") or f"v{overwrite_version}"
                db.execute(
                    text(
                        f"UPDATE saved_schedule_versions "
                        f"SET version_name=:vn, schedule_data={json_cast('d')}, published=:p, updated_at={sql_now()} "
                        f"WHERE year=:y AND month=:m AND version=:v"
                    ),
                    {"vn": final_name, "d": json.dumps(sched_by_name), "p": publish, "y": year, "m": month, "v": overwrite_version}
                )
                next_ver = overwrite_version
            else:
                row = db.execute(text("SELECT COALESCE(MAX(version),0) AS maxv FROM saved_schedule_versions WHERE year=:y AND month=:m"),
                                 {"y": year, "m": month}).mappings().fetchone()
                next_ver = (row['maxv'] or 0) + 1
                final_name = version_name or f"v{next_ver}"
                db.execute(
                    text(f"INSERT INTO saved_schedule_versions (year, month, version, version_name, schedule_data, published) "
                         f"VALUES (:y, :m, :v, :vn, {json_cast('d')}, :p)"),
                    {"y": year, "m": month, "v": next_ver, "vn": final_name, "d": json.dumps(sched_by_name), "p": publish}
                )
            if publish:
                db.execute(
                    text(f"UPDATE saved_schedule_versions SET published = {sql_false()} "
                         "WHERE year=:y AND month=:m AND version <> :v"),
                    {"y": year, "m": month, "v": next_ver}
                )

        action = "Updated" if overwrite_version is not None else "Saved"
        flash(f"{action} version v{next_ver} ({final_name}){' and published' if publish else ''}.", 'success')
        return redirect(url_for('edit_publish', year=year, month=month))

    #############################################
    # Start, poll, cancel endpoints
    #############################################

    # --- Route to start a new solve job ---
    @app.route('/start_solve', methods=['POST'])
    @basic_auth.required
    def start_solve():
        data = request.get_json()
        year, month = int(data['year']), int(data['month'])
        time_limit_seconds = int(data.get('time_limit_seconds', 30))
        allow_empty = bool(data.get('allow_empty', False))
        prior_last_two = data.get('prior_last_two') or {}

        # Build days list
        days = [
            datetime.date(year, month, d).isoformat()
            for d in range(1, calendar.monthrange(year, month)[1] + 1)
        ]

        # Load previous month’s schedule
        if month == 1:
            prev_year, prev_month = year - 1, 12
        else:
            prev_year, prev_month = year, month - 1

        db = get_db()
        # --- Fetch as a mapping so we can use column names ---
        stmt = text(
            "SELECT schedule_data FROM saved_schedule "
            "WHERE year = :y AND month = :m"
        )
        result = db.execute(stmt, {"y": prev_year, "m": prev_month})
        prev_row = result.mappings().fetchone()
        # Ignore saved previous schedules; rely only on user-provided prior_last_two
        prev_schedule = None


        # Compute HK public holidays
        hk_h = holidays.HK(years=[year])
        public_holidays = {
            d.isoformat() for d in hk_h
            if d.year == year and d.month == month
        }
        # ---- Load preassignments for the current month ----
        stmt_pre = text(
            "SELECT preassignment_data FROM preassignments WHERE year = :year AND month = :month"
        )
        result_pre = db.execute(stmt_pre, {"year": year, "month": month})
        row_pre = result_pre.mappings().fetchone()
        if row_pre and row_pre['preassignment_data'] is not None and row_pre['preassignment_data'] != "":
            raw_pre = row_pre['preassignment_data']
            try:
                preassignments = raw_pre if isinstance(raw_pre, dict) else json.loads(raw_pre)
            except Exception as e:
                print("Error decoding preassignments:", e)
                preassignments = {}
        else:
            preassignments = {}

        surgeons = get_all_surgeons()
        # Half-year horizon prior counts and credits
        cfg = get_global_config()
        enable_horizon = str(cfg.get('enable_horizon_fairness','0')) == '1'
        horizon_prior = None
        if enable_horizon:
            from helper import get_horizon_prior_levels_and_credit
            horizon_prior = get_horizon_prior_levels_and_credit(year, month, surgeons)
        # Build prev_schedule only from prior_last_two for carry-over spacing
        if prior_last_two:
            # Build date strings for last two days of the previous month
            import calendar as _cal
            last_day = _cal.monthrange(year if month>1 else year-1, month-1 if month>1 else 12)[1]
            p_year = year if month>1 else year-1
            p_month = month-1 if month>1 else 12
            day_m2 = datetime.date(p_year, p_month, max(1, last_day-1)).isoformat()
            day_m1 = datetime.date(p_year, p_month, last_day).isoformat()
            # ensure dict
            if prev_schedule is None:
                prev_schedule = {}
            # convert ids→names for the solver's prev_schedule format
            id_to_name = {s['id']: s['name'] for s in surgeons}
            m2 = prior_last_two.get('m2') or {}
            m1 = prior_last_two.get('m1') or {}
            if m2:
                prev_schedule.setdefault(day_m2, {})
                for lev, sid in m2.items():
                    name = id_to_name.get(int(sid))
                    if name:
                        prev_schedule[day_m2][lev] = name
            if m1:
                prev_schedule.setdefault(day_m1, {})
                for lev, sid in m1.items():
                    name = id_to_name.get(int(sid))
                    if name:
                        prev_schedule[day_m1][lev] = name

        job_id = uuid.uuid4().hex
        threading.Thread(
            target=run_solver_job,
            args=(job_id, days, surgeons, prev_schedule, public_holidays, preassignments, time_limit_seconds, allow_empty, horizon_prior),
            daemon=True
        ).start()
        
        return jsonify({'job_id': job_id})

    # --- Route to poll status of a job ---
    @app.route('/solve_status/<job_id>')
    @basic_auth.required
    def solve_status(job_id):
        job = solve_jobs.get(job_id)
        if not job:
            return jsonify({'status': 'unknown'}), 404
        resp = {
            'status': job['status'],
            'best':   job['best'],
            'public_holidays': job.get('public_holidays', []),
        }
        if job['status'] == 'done':
            resp['solution'] = job['solution']
            resp['solver_mode'] = job.get('solver_mode')
        elif job['status'] == 'failed':
            resp['errors'] = job.get('solution', {}).get('errors') if isinstance(job.get('solution'), dict) else None
            resp['analysis'] = job.get('solution', {}).get('analysis') if isinstance(job.get('solution'), dict) else None
            resp['solver_mode'] = job.get('solver_mode')
        return jsonify(resp)

    # --- Route to cancel a running job ---
    @app.route('/cancel_solve/<job_id>', methods=['POST'])
    @basic_auth.required
    def cancel_solve(job_id):
        job = solve_jobs.get(job_id)
        if job and job['status'] == 'running':
            job['cancel'] = True
            return jsonify({'status': 'cancelling'})
        return jsonify({'status': 'not_found'}), 404



#############################################
# Preassignment
#############################################
    @app.route('/preassignment', methods=['GET', 'POST'])
    @basic_auth.required
    def preassignment():
        import calendar, datetime
        from helper import get_all_surgeons, parse_call_levels
        db = get_db()
        
        if request.method == 'POST':
            # Process form submission
            preassignments = {}
            # Each field is named "preassignment_<day>_<level>"
            for key, value in request.form.items():
                if key.startswith("preassignment_"):
                    parts = key.split("_")
                    # parts[1] = day (ISO date string), parts[2] = level
                    day = parts[1]
                    level = parts[2]
                    if value:
                        preassignments.setdefault(day, {})[level] = int(value)
                    else:
                        preassignments.setdefault(day, {})[level] = None

            # Save to database – if record exists update; otherwise insert.
            year = request.args.get('year', None, type=int)
            month = request.args.get('month', None, type=int)
            if not year or not month:
                today = datetime.date.today()
                year, month = today.year, today.month

            with db.begin():
                row = db.execute(
                    text("SELECT id FROM preassignments WHERE year = :year AND month = :month"),
                    {"year": year, "month": month}
                ).fetchone()
                if row:
                    row_dict = row._mapping  # Access the row as a mapping object
                    db.execute(
                        text(f"UPDATE preassignments SET preassignment_data = {json_cast('data')}, date_updated = {sql_now()} WHERE id = :id"),
                        {"data": json.dumps(preassignments), "id": row_dict["id"]}
                    )
                else:
                    db.execute(
                        text(f"INSERT INTO preassignments (year, month, preassignment_data) VALUES (:year, :month, {json_cast('data')})"),
                        {"year": year, "month": month, "data": json.dumps(preassignments)}
                    )
                flash("Preassignments updated successfully!", "success")
                return redirect(url_for('preassignment', year=year, month=month))
        else:
            # GET: load the desired year/month and build a table.
            year = request.args.get('year', None, type=int)
            month = request.args.get('month', None, type=int)
            if not year or not month:
                today = datetime.date.today()
                year, month = today.year, today.month
            days = [
                datetime.date(year, month, d).isoformat()
                for d in range(1, calendar.monthrange(year, month)[1] + 1)
            ]
            # Build candidate lists for each level similar to scheduler.py logic
            surgeons = get_all_surgeons()
            candidates = {}
            for level in ["1A", "1B", "2A", "2B", "3", "4"]:
                candidate_options = [s for s in surgeons if level in parse_call_levels(s.get("call_levels", ""))]
                candidate_options.sort(key=lambda s: s["name"])
                candidates[level] = candidate_options
            # Load any existing preassignments from the database
            row = db.execute(
                text("SELECT preassignment_data FROM preassignments WHERE year = :year AND month = :month"),
                {"year": year, "month": month}
            ).fetchone()
            if row:
                row_dict = row._mapping
                preassignments = coerce_json_column(row_dict["preassignment_data"]) or {}
            else:
                preassignments = {}
            return render_template(
                "preassignment.html",
                year=year,
                month=month,
                days=days,
                levels=["1A", "1B", "2A", "2B", "3", "4"],
                candidates=candidates,
                preassignments=preassignments
            )
        

###############################################
# End of App
################################################

    return app

#############################################
# Run the App
#############################################
if __name__=="__main__":
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    create_app().run(host="0.0.0.0", port=int(os.environ.get("PORT",5000)), debug=debug_mode)